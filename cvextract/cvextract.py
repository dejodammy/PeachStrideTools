#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cvextract - pull names, emails and phone numbers out of a folder of CVs.

Runs entirely on this machine. No network, no API key, no model. The same input
always produces the same output, and every decision below is a line you can read.

    python cvextract.py <folder>
    python cvextract.py <folder> -o contacts.xlsx
    python cvextract.py <folder> --master contacts_master.xlsx
    python cvextract.py <folder> --watch

WHAT IT IS HONEST ABOUT
    It gets most rows right on its own and marks the rest for you. A CV that puts
    the candidate's email in a Word text box will otherwise hand you the REFEREE's
    address under the candidate's name - no blank, no error, just a confident
    wrong answer. Rows it is unsure about land on a "Needs Review" sheet rather
    than being quietly guessed at.

REQUIREMENTS
    pip install pymupdf python-docx openpyxl pillow lxml
    OCR for scanned CVs uses the Windows on-device engine (Windows 10/11).
    Legacy .doc needs Word or LibreOffice installed; without either it is skipped
    and reported, never silently dropped.
"""
from __future__ import annotations

import argparse
import csv
import glob
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile

SUPPORTED = {'.pdf', '.docx', '.doc'}
SKIP_EXT = {'.xlsx', '.xls', '.csv', '.txt', '.zip', '.jpg', '.png', '.jpeg'}

# ---------------------------------------------------------------------------
# Patterns
# ---------------------------------------------------------------------------
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

# Anything that opens a block of OTHER people's contact details. Matched on its
# own line, and tolerant of the letter-spacing designed CVs like to use
# ("R E F E R E N C E S").
REF_RE = re.compile(
    r"^[\s\W]*(?:REFEREE?S?|REFERENCES?|RECOMMENDATIONS?|"
    r"R\s+E\s+F\s+E\s+R\s+E\s+N\s+C\s+E\s+S)\s*:?[\s\W]*$", re.I | re.M)

# "Name: Adebusola Seyi Moses" in a PERSONAL DATA block.
NAME_FIELD_RE = re.compile(
    r"^[\s\W]*(?:FULL\s+)?NAMES?\s*[:\-]\s*(.+)$", re.I | re.M)

# A label glued to the address because Word auto-linked the whole line:
# "Email-asuwaniemmanuel@mail.com".
LABEL_RE = re.compile(r"^(?:e[\s\-]?mail|mail|address|contact)[\s\-.:]+", re.I)

# Digits plus the punctuation Nigerian CVs sprinkle through them: +234, (0), -.
PHONE_RUN_RE = re.compile(r"(?:\+?234|0)[\d\s\-().]{8,24}\d")

# Fallback for candidates based abroad: a run that opens with an explicit
# international marker (+ or the 00 access code) rather than Nigeria's own
# "0" or "234" — those are ambiguous with plain local numbers on their own,
# so this only ever runs once the Nigerian pass above has come up empty.
INTL_PHONE_RUN_RE = re.compile(r"(?:\+|00)\d[\d\s\-().]{6,20}\d")

COMMON_DOMAINS = ['gmail.com', 'yahoo.com', 'outlook.com', 'hotmail.com',
                  'yahoo.co.uk', 'googlemail.com', 'icloud.com', 'live.com']

HEADINGS = {
    'contact', 'contacts', 'curriculum vitae', 'resume', 'cv', 'personal data',
    'personal details', 'personal information', 'profile', 'summary', 'about me',
    'professional summary', 'career objective', 'objective', 'career objectives',
    'skills', 'skill highlight', 'education', 'experience', 'work experience',
    'references', 'referees', 'phone', 'email', 'address', 'bio-data', 'bio data',
    'professional profile', 'core competencies', 'languages', 'hobbies',
    'work history', 'employment history', 'qualifications', 'personal statement',
    'declaration', 'interests', 'achievements', 'certifications', 'soft skills',
    'hard skills', 'technical skills', 'key achievements', 'career summary',
}
# Words that mean a line is an address, not a name.
STOP_TOKENS = {
    'street', 'road', 'avenue', 'close', 'crescent', 'estate', 'quarters',
    'junction', 'behind', 'opposite', 'block', 'plot', 'house', 'village',
    'nigeria', 'nigerian', 'state', 'fct', 'abuja', 'lagos', 'kano', 'jos',
    'university', 'polytechnic', 'college', 'school', 'academy', 'institute',
    'limited', 'ltd', 'company', 'bank', 'plc', 'hotel', 'diploma', 'degree',
    'certificate', 'national', 'senior', 'secondary', 'primary', 'obtained',
    'department', 'manager', 'officer', 'assistant', 'intern', 'available',
    'request', 'present', 'current',
}


# ---------------------------------------------------------------------------
# OCR - Windows on-device engine, written out on first use
# ---------------------------------------------------------------------------
OCR_PS1 = r'''
param([string]$ImagePath)
$ErrorActionPreference = 'Stop'
Add-Type -AssemblyName System.Runtime.WindowsRuntime | Out-Null
$asTaskGeneric = ([System.WindowsRuntimeSystemExtensions].GetMethods() | Where-Object {
    $_.Name -eq 'AsTask' -and $_.GetParameters().Count -eq 1 -and
    $_.GetParameters()[0].ParameterType.Name -eq 'IAsyncOperation`1' })[0]
function Await($WinRtTask, $ResultType) {
    $asTask = $asTaskGeneric.MakeGenericMethod($ResultType)
    $netTask = $asTask.Invoke($null, @($WinRtTask))
    $netTask.Wait(-1) | Out-Null
    $netTask.Result
}
[Windows.Storage.StorageFile,Windows.Storage,ContentType=WindowsRuntime] | Out-Null
[Windows.Graphics.Imaging.BitmapDecoder,Windows.Graphics.Imaging,ContentType=WindowsRuntime] | Out-Null
[Windows.Media.Ocr.OcrEngine,Windows.Foundation,ContentType=WindowsRuntime] | Out-Null
$full = (Resolve-Path $ImagePath).Path
$file = Await ([Windows.Storage.StorageFile]::GetFileFromPathAsync($full)) ([Windows.Storage.StorageFile])
$stream = Await ($file.OpenAsync([Windows.Storage.FileAccessMode]::Read)) ([Windows.Storage.Streams.IRandomAccessStream])
$decoder = Await ([Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($stream)) ([Windows.Graphics.Imaging.BitmapDecoder])
$bitmap = Await ($decoder.GetSoftwareBitmapAsync()) ([Windows.Graphics.Imaging.SoftwareBitmap])
$engine = [Windows.Media.Ocr.OcrEngine]::TryCreateFromUserProfileLanguages()
if ($null -eq $engine) { Write-Output "[[NO_OCR_ENGINE]]"; exit 2 }
$result = Await ($engine.RecognizeAsync($bitmap)) ([Windows.Media.Ocr.OcrResult])
foreach ($line in $result.Lines) { Write-Output $line.Text }
'''

_ocr_script = None


def ocr_script_path():
    global _ocr_script
    if _ocr_script is None:
        d = tempfile.mkdtemp(prefix='cvextract_')
        _ocr_script = os.path.join(d, 'ocr.ps1')
        with open(_ocr_script, 'w', encoding='utf-8') as f:
            f.write(OCR_PS1)
    return _ocr_script


def ocr_image(png_path):
    if sys.platform != 'win32':
        return ''
    try:
        r = subprocess.run(
            ['powershell.exe', '-NoProfile', '-ExecutionPolicy', 'Bypass',
             '-File', ocr_script_path(), '-ImagePath', png_path],
            capture_output=True, timeout=120)
        return r.stdout.decode('utf-8', 'replace')
    except Exception:
        return ''


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------
W = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
DOCX_PARTS = re.compile(
    r'^word/(document|header\d*|footer\d*|footnotes|endnotes)\.xml$')


def docx_text(path):
    """Walk the WordprocessingML in document order.

    python-docx's .paragraphs never descends into w:txbxContent, so anything the
    author put in a TEXT BOX is invisible to it - and on a CV that is often the
    whole contact block. Dropping it does not leave a blank, it leaves the
    referee's address as the only one in the file. Pre-order traversal picks up
    text boxes, tables, headers and footers without double-counting nesting.

    w:instrText is skipped on purpose: it holds field codes such as
    HYPERLINK "mailto:..." whose target is often a stale address the author
    edited away from, and is not what a reader of the CV sees.
    """
    from lxml import etree
    z = zipfile.ZipFile(path)
    parts = sorted((n for n in z.namelist() if DOCX_PARTS.match(n)),
                   key=lambda n: (n != 'word/document.xml', n))
    out = []
    for part in parts:
        try:
            root = etree.fromstring(z.read(part))
        except Exception:
            continue
        for el in root.iter():
            if el.tag in (W + 'p', W + 'br', W + 'tr'):
                out.append('\n')
            elif el.tag == W + 'tab':
                out.append('\t')
            elif el.tag == W + 't':
                out.append(el.text or '')
    return re.sub(r'\n[ \t]*\n+', '\n', ''.join(out))


def doc_text(path):
    """Legacy .doc via Word COM, falling back to LibreOffice."""
    if sys.platform == 'win32':
        ps = (
            "$w=New-Object -ComObject Word.Application; $w.Visible=$false; "
            "$d=$w.Documents.Open('%s',$false,$true); $t=$d.Content.Text; "
            "$d.Close(0); $w.Quit(); [Console]::Out.Write($t)"
            % os.path.abspath(path).replace("'", "''"))
        try:
            r = subprocess.run(['powershell.exe', '-NoProfile', '-Command', ps],
                               capture_output=True, timeout=120)
            t = r.stdout.decode('utf-8', 'replace')
            if t.strip():
                return t.replace('\r', '\n')
        except Exception:
            pass
    soffice = shutil.which('soffice') or shutil.which('libreoffice')
    if soffice:
        d = tempfile.mkdtemp()
        try:
            subprocess.run([soffice, '--headless', '--convert-to', 'txt:Text',
                            '--outdir', d, os.path.abspath(path)],
                           capture_output=True, timeout=180)
            for f in glob.glob(os.path.join(d, '*.txt')):
                return open(f, encoding='utf-8', errors='replace').read()
        except Exception:
            pass
    return ''


def page_image(page, doc, tmp, i):
    """Prefer the embedded scan at its native resolution. Rendering a low-res
    scan at high DPI only interpolates - it adds no detail the OCR can use."""
    import pymupdf
    imgs = page.get_images(full=True)
    if len(imgs) == 1:
        pix = pymupdf.Pixmap(doc, imgs[0][0])
        if pix.n > 3:
            pix = pymupdf.Pixmap(pymupdf.csRGB, pix)
        png = os.path.join(tmp, 'e%d.png' % i)
        pix.save(png)
        return png
    png = os.path.join(tmp, 'r%d.png' % i)
    page.get_pixmap(dpi=200).save(png)
    return png


def pdf_ocr(path, max_pages=3):
    import pymupdf
    from PIL import Image, ImageFilter, ImageOps
    out = []
    tmp = tempfile.mkdtemp()
    with pymupdf.open(path) as d:
        for i, page in enumerate(d):
            if i >= max_pages:
                break
            png = page_image(page, d, tmp, i)
            im = Image.open(png).convert('L')
            if im.width < 2200:
                f = max(2, round(2200 / im.width))
                im = im.resize((im.width * f, im.height * f), Image.LANCZOS)
                im = ImageOps.autocontrast(im)
                im = im.filter(ImageFilter.SHARPEN)
            im.save(png)
            out.append(ocr_image(png))
    return "\n".join(out)


def read_document(path):
    """-> (text, how). 'how' feeds the triage layer."""
    import pymupdf
    ext = os.path.splitext(path)[1].lower()
    if ext == '.docx':
        return docx_text(path), 'docx'
    if ext == '.doc':
        t = doc_text(path)
        return t, ('doc' if t.strip() else 'doc-failed')
    with pymupdf.open(path) as d:
        txt = "\n".join(p.get_text() for p in d)
    if len(txt.strip()) < 200:
        return (txt + "\n" + pdf_ocr(path)), 'pdf+ocr'
    return txt, 'pdf'


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def local_part(email):
    return re.sub(r'[^a-z0-9]', '', email.split('@')[0].lower())


def alpha(s):
    return re.sub(r'[^a-z]', '', (s or '').lower())


def overlap(a, b):
    """Longest run of a appearing in b, as a fraction of a."""
    best = 0
    for i in range(len(a)):
        for j in range(i + best + 1, len(a) + 1):
            if a[i:j] in b:
                best = max(best, j - i)
            else:
                break
    return best / max(len(a), 1)


def edit_distance(a, b, cap=3):
    if abs(len(a) - len(b)) > cap:
        return cap + 1
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def body_before_referees(text):
    m = REF_RE.search(text)
    return text[:m.start()] if m else text


# ---------------------------------------------------------------------------
# Field extraction
# ---------------------------------------------------------------------------
def find_emails(text, filename):
    """-> (chosen, all_candidates, flags)."""
    flags = []
    head = body_before_referees(text)
    raw = EMAIL_RE.findall(head)
    if not raw:
        raw = EMAIL_RE.findall(text)
        if raw:
            flags.append('EMAIL_ONLY_IN_REFEREE_BLOCK')
    if not raw:
        return None, [], ['NO_EMAIL']

    cands, seen = [], set()
    for e in raw:
        stripped = LABEL_RE.sub('', e)
        if stripped != e and EMAIL_RE.fullmatch(stripped):
            if 'EMAIL_LABEL_STRIPPED' not in flags:
                flags.append('EMAIL_LABEL_STRIPPED')
            e = stripped
        if e.lower() not in seen:
            seen.add(e.lower())
            cands.append(e)

    stem = alpha(os.path.splitext(filename)[0])
    cands.sort(key=lambda e: -overlap(local_part(e), stem))
    chosen = cands[0]

    if len(cands) > 1:
        flags.append('MULTI_EMAIL')
    dom = chosen.split('@')[-1].lower()
    if dom not in COMMON_DOMAINS:
        for good in COMMON_DOMAINS:
            if 0 < edit_distance(dom, good) <= 2:
                flags.append('DOMAIN_NEAR_' + good)
                break
    return chosen, cands, flags


def normalise_phone(d):
    if d.startswith('234'):
        d = d[3:]
    if len(d) == 10 and d[0] in '789':
        d = '0' + d
    if len(d) == 11 and d[0] == '0' and d[1] in '789':
        return d
    return None


def split_run(d):
    """Two numbers written '0816... / 0813...' become one digit run once the
    separators are stripped. Peel valid numbers off the front rather than
    treating the whole run as a single (invalid) number."""
    out = []
    while len(d) >= 10:
        if d.startswith('234') and len(d) >= 14 and d[3] == '0':
            take = 14
        elif d.startswith('234') and len(d) >= 13:
            take = 13
        elif d[0] == '0' and len(d) >= 11:
            take = 11
        elif d[0] in '789' and len(d) >= 10:
            take = 10
        else:
            d = d[1:]
            continue
        n = normalise_phone(d[:take])
        if n:
            out.append(n)
            d = d[take:]
        else:
            d = d[1:]
    return out


def normalise_foreign_phone(raw):
    """A matched international run -> '+<countrycode><number>', or None if the
    digit count isn't plausible for a real phone number (E.164 allows up to 15
    digits after the country code marker)."""
    digits = re.sub(r'\D', '', raw)
    if raw.strip().startswith('00'):
        digits = digits[2:]  # '00' is the access code, not part of the number
    if not (8 <= len(digits) <= 15):
        return None
    return '+' + digits


def find_phones(text, limit=2):
    """-> (phones, flags). Never reads below a REFEREES heading. A Nigerian
    mobile is always preferred; a foreign number is only reported when the CV
    has no Nigerian number at all, and is flagged so it gets a second look —
    the pattern that catches it is far looser than the Nigerian one."""
    head = body_before_referees(text)
    out = []
    for raw in PHONE_RUN_RE.findall(head):
        digits = re.sub(r'\D', '', raw)
        # A run starting '00' opens with the international access code, not a
        # Nigerian trunk prefix — leave it for the foreign pass below rather
        # than let split_run peel a plausible-looking Nigerian number out of
        # someone else's country code.
        if digits.startswith('00'):
            continue
        for n in split_run(digits):
            if n not in out:
                out.append(n)
    if out:
        return out[:limit], []

    foreign = []
    for raw in INTL_PHONE_RUN_RE.findall(head):
        n = normalise_foreign_phone(raw)
        if n and n not in foreign:
            foreign.append(n)
    if foreign:
        return foreign[:limit], ['FOREIGN_PHONE']

    return [], ['NO_PHONE']


def looks_like_name(line):
    s = line.strip().strip('|').strip()
    if not (4 <= len(s) <= 45) or '@' in s or ':' in s:
        return False
    if re.search(r'\d', s) or s.lower().strip('.') in HEADINGS:
        return False
    toks = re.findall(r"[A-Za-z][A-Za-z'\-.]*", s)
    if not (2 <= len(toks) <= 5):
        return False
    if len(toks) != len(s.replace(',', ' ').split()):
        return False
    if any(t.lower().strip('.') in STOP_TOKENS for t in toks):
        return False
    return sum(1 for t in toks if t[0].isupper()) >= len(toks) - 1


def find_name(text, email, filename):
    """-> (name, score, flags). Score drives the NAME_LOW_CONFIDENCE flag."""
    flags = []

    # A "Name:" field in a PERSONAL DATA block beats any line-shape guess.
    for m in NAME_FIELD_RE.finditer(body_before_referees(text)):
        cand = m.group(1).strip().strip('.').strip()
        cand = re.split(r'\s{3,}|\t', cand)[0].strip()
        if cand and looks_like_name(cand):
            return cand, 3.0, ['NAME_FROM_FIELD']

    loc = local_part(email) if email else ''
    stem = alpha(os.path.splitext(filename)[0])
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    best, best_score = None, -1.0
    for i, l in enumerate(lines[:40]):
        if not looks_like_name(l):
            continue
        flat = alpha(l)
        score = (3.0 * overlap(flat, loc) + 1.5 * overlap(flat, stem)
                 + max(0.0, 1.2 - i * 0.12) + (0.5 if l.isupper() else 0.0))
        if score > best_score:
            best, best_score = l.strip(), score

    # A name split over two lines ("AHUCHE" / "LINDA CHIGOZIE") gets picked up as
    # only the second half. Join the neighbour when the EMAIL vouches for it: the
    # neighbour's letters must appear in the local part and not already be covered
    # by the line we chose. Requiring that corroboration is what keeps this from
    # firing on ordinary CVs, where the line above a name is a heading or a job
    # title that the address says nothing about.
    if best is not None and loc:
        bi = lines.index(best) if best in lines else -1
        for nb_i in (bi - 1, bi + 1):
            if bi < 0 or not (0 <= nb_i < len(lines)):
                continue
            nb_line = lines[nb_i]
            # The line next to a name is very often the address line itself, whose
            # tokens match the local part trivially and mean nothing. Skip those.
            if '@' in nb_line or re.search(r'\d', nb_line):
                continue
            # Match per token, not per line: a designed CV can leave the stray tail
            # of another frame on the same line as the name ("Igbo AHUCHE").
            toks = re.findall(r"[A-Za-z][A-Za-z'\-]*", nb_line)
            if not toks or len(toks) > 4:
                continue
            hits = [t for t in toks
                    # strictly shorter than the local part: a token that accounts for
                    # the whole address is the address, not the missing half of a name
                    if 4 <= len(alpha(t)) < len(loc) and alpha(t) in loc
                    and alpha(t) not in alpha(best)
                    and t.lower().strip('.') not in HEADINGS
                    and t.lower().strip('.') not in STOP_TOKENS]
            if not hits:
                continue
            take = hits[-1] if nb_i < bi else hits[0]
            best = (take + ' ' + best) if nb_i < bi else (best + ' ' + take)
            flags.append('NAME_JOINED_FROM_TWO_LINES')
            break

    if best is None:
        guess = re.sub(r'[_\-.]+', ' ', os.path.splitext(filename)[0])
        guess = re.sub(r'(?i)\b(cv|resume|updated|new|final|professional|\d+)\b', ' ', guess)
        guess = re.sub(r'[^A-Za-z ]', ' ', guess)
        guess = ' '.join(guess.split())
        return (titlecase(guess) if guess else ''), 0.0, ['NAME_FROM_FILENAME']

    if best_score < 1.6:
        flags.append('NAME_LOW_CONFIDENCE')
    return best, best_score, flags


def titlecase(s):
    if not s:
        return s
    return ' '.join(w if re.search(r'[a-z]', w) else w.title() for w in s.split())


# ---------------------------------------------------------------------------
# Per-file pipeline
# ---------------------------------------------------------------------------
def process_file(path):
    fn = os.path.basename(path)
    rec = {'file': fn, 'name': '', 'email': '', 'phones': '', 'how': '',
           'flags': [], 'candidates': ''}
    ext = os.path.splitext(fn)[1].lower()
    if ext not in SUPPORTED:
        rec['how'] = 'skipped'
        rec['flags'] = ['UNSUPPORTED_FORMAT']
        return rec
    try:
        text, how = read_document(path)
    except Exception as exc:
        rec['how'] = 'error'
        rec['flags'] = ['READ_FAILED: %s' % type(exc).__name__]
        return rec

    rec['how'] = how
    if how == 'doc-failed':
        rec['flags'] = ['DOC_NEEDS_WORD_OR_LIBREOFFICE']
        return rec
    if not text.strip():
        rec['flags'] = ['NO_TEXT_FOUND']
        return rec

    email, cands, ef = find_emails(text, fn)
    phones, pf = find_phones(text)
    name, score, nf = find_name(text, email, fn)

    rec['email'] = email or ''
    rec['name'] = titlecase(name) if name else ''
    rec['phones'] = ' / '.join(phones)
    rec['candidates'] = '; '.join(cands[1:]) if len(cands) > 1 else ''
    rec['flags'] = ef + pf + nf + (['OCR_USED'] if how == 'pdf+ocr' else [])
    return rec


def needs_review(rec):
    """Which flags mean a person should look. Deliberately conservative: a row
    that is silently wrong costs far more than a row you glance at."""
    hard = ('NO_EMAIL', 'MULTI_EMAIL', 'OCR_USED', 'NAME_LOW_CONFIDENCE',
            'NAME_FROM_FILENAME', 'UNSUPPORTED_FORMAT', 'NO_TEXT_FOUND',
            'READ_FAILED', 'DOC_NEEDS_WORD', 'DOMAIN_LOOKS_LIKE_TYPO',
            'EMAIL_LABEL_STRIPPED', 'EMAIL_ONLY_IN_REFEREE_BLOCK',
            'DUPLICATE_IN_BATCH', 'NO_PHONE', 'FOREIGN_PHONE',
            'NAME_JOINED_FROM_TWO_LINES')
    return any(f.startswith(h) for f in rec['flags'] for h in hard)


FLAG_HELP = {
    'NO_EMAIL': 'No address found anywhere in the document.',
    'MULTI_EMAIL': 'More than one address outside the referee block; the best name match was used. Others are in "Other Emails".',
    'OCR_USED': 'No text layer - read by OCR from the page image. Check the characters.',
    'NAME_LOW_CONFIDENCE': 'No line looked clearly like a name; the best guess was used.',
    'NAME_FROM_FILENAME': 'No name found in the document at all; taken from the file name.',
    'NAME_FROM_FIELD': 'Name read from a "Name:" field rather than a heading. Usually reliable.',
    'NAME_JOINED_FROM_TWO_LINES': 'The name was split across two lines and rejoined, because the email address contained both halves. Worth a glance.',
    'UNSUPPORTED_FORMAT': 'Not a CV format this tool reads. Nothing was extracted.',
    'NO_TEXT_FOUND': 'File opened but contained no readable text.',
    'DOC_NEEDS_WORD_OR_LIBREOFFICE': 'Legacy .doc needs Word or LibreOffice installed to read.',
    'EMAIL_LABEL_STRIPPED': 'A label was glued to the address (e.g. "Email-name@x.com"). The label was removed - confirm the result.',
    'EMAIL_ONLY_IN_REFEREE_BLOCK': 'The only address sits under a REFEREES heading. It may belong to the referee, not the candidate.',
    'DUPLICATE_IN_BATCH': 'This address appears on more than one CV in this batch.',
    'DOMAIN_NEAR': 'The domain is 1-2 letters away from a common provider (gmaill.com, gamil.com). It may be a typo on the CV - or a real, less common provider such as mail.com. Confirm before sending.',
    'NO_PHONE': 'No phone number, Nigerian or foreign, found outside the referee block.',
    'FOREIGN_PHONE': 'No Nigerian mobile on this CV; a non-Nigerian number was used instead. The pattern for these is looser, so check the country code and digit count.',
}


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
def write_workbook(records, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    HEAD = PatternFill('solid', fgColor='1F3864')
    HFONT = Font(color='FFFFFF', bold=True)
    WARN = PatternFill('solid', fgColor='FFF2CC')

    wb = Workbook()
    wb.remove(wb.active)

    def sheet(title, header, rows, widths, wrap_from=99):
        ws = wb.create_sheet(title)
        ws.append(header)
        for r in rows:
            ws.append(r)
        for c in ws[1]:
            c.fill, c.font = HEAD, HFONT
        ws.freeze_panes = 'A2'
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical='top',
                                        wrap_text=(c.column >= wrap_from))
        return ws

    cols = ['Name', 'Email', 'Phone', 'Source File', 'Read As',
            'Other Emails', 'Flags']
    widths = [26, 32, 26, 42, 12, 34, 50]

    ok = [r for r in records if not needs_review(r)]
    review = [r for r in records if needs_review(r)]
    row = lambda r: [r['name'], r['email'], r['phones'], r['file'], r['how'],
                     r['candidates'], ', '.join(r['flags'])]

    sheet('Contacts', cols, [row(r) for r in records], widths, 7)
    ws = wb['Contacts']
    for wsrow in ws.iter_rows(min_row=2):
        if wsrow[6].value:
            for c in wsrow:
                c.fill = WARN

    sheet('Needs Review', cols, [row(r) for r in review], widths, 7)
    sheet('Clean', cols, [row(r) for r in ok], widths, 7)

    used = sorted({f.split(':')[0] for r in records for f in r['flags']})
    helprows = []
    for f in used:
        key = next((k for k in FLAG_HELP if f.startswith(k)), None)
        helprows.append([f, FLAG_HELP.get(key, 'See the source for this flag.')])
    if not helprows:
        helprows = [['(none)', 'Nothing was flagged in this batch.']]
    sheet('Flag Guide', ['Flag', 'What it means'], helprows, [40, 100], 2)

    wb.save(out_path)
    return len(records), len(ok), len(review)


def merge_into_master(records, master_path):
    """Append new people to an existing master workbook.

    Policy: existing rows are preserved verbatim. Nothing is renamed, reformatted
    or repaired - a malformed row is flagged, never fixed, because guessing at
    somebody's contact address is fabrication dressed as a correction.
    """
    from openpyxl import load_workbook

    wb = load_workbook(master_path)
    # Find the sheet that actually holds contacts. A workbook may lead with a
    # Read Me or a cover sheet, so pick by header rather than by position.
    ws = header = None
    for cand in wb.worksheets:
        h = [str(c.value or '').strip().lower() for c in cand[1]]
        if 'email' in h:
            ws, header = cand, h
            break
    if ws is None:
        raise SystemExit('! no sheet in %s has an "Email" column' % master_path)
    ci_email = header.index('email')
    ci_name = header.index('name') if 'name' in header else 0
    ci_phone = header.index('phone') if 'phone' in header else None

    have = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if ci_email < len(r) and r[ci_email]:
            have.add(str(r[ci_email]).strip().lower())

    added = 0
    for rec in records:
        em = rec['email'].strip().lower()
        if not em or em in have:
            continue
        have.add(em)
        row = [''] * len(header)
        row[ci_name] = rec['name']
        row[ci_email] = rec['email']
        if ci_phone is not None:
            row[ci_phone] = rec['phones']
        for label, idx in (('status', None), ('source', None), ('notes', None)):
            if label in header:
                idx = header.index(label)
                row[idx] = {'status': 'New from CV', 'source': rec['file'],
                            'notes': ', '.join(rec['flags'])}[label]
        ws.append(row)
        added += 1
    wb.save(master_path)
    return added


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def collect(folder):
    out = []
    for fn in sorted(os.listdir(folder)):
        p = os.path.join(folder, fn)
        if not os.path.isfile(p):
            continue
        ext = os.path.splitext(fn)[1].lower()
        if ext in SKIP_EXT:
            continue
        out.append(p)
    return out


def run_once(folder, out_path, master, quiet=False):
    files = collect(folder)
    if not files:
        print('No CV files in %s' % folder)
        return []
    records = []
    for i, p in enumerate(files, 1):
        rec = process_file(p)
        records.append(rec)
        if not quiet:
            mark = '!' if needs_review(rec) else ' '
            print('%s [%2d/%2d] %-42s %-24s %-30s %s'
                  % (mark, i, len(files), rec['file'][:42], rec['name'][:24],
                     rec['email'][:30], rec['phones']))

    seen = {}
    for r in records:
        if r['email']:
            seen.setdefault(r['email'].lower(), []).append(r)
    for em, group in seen.items():
        if len(group) > 1:
            for r in group:
                r['flags'].append('DUPLICATE_IN_BATCH')

    total, ok, review = write_workbook(records, out_path)
    print('\n%d CVs | %d clean | %d need review -> %s'
          % (total, ok, review, os.path.abspath(out_path)))
    if master:
        if not os.path.exists(master):
            print('! master not found: %s' % master)
        else:
            n = merge_into_master([r for r in records if not needs_review(r)], master)
            print('%d new people appended to %s (rows needing review were held back)'
                  % (n, master))
    return records


def main():
    ap = argparse.ArgumentParser(
        description='Extract names, emails and phone numbers from a folder of CVs. '
                    'Runs entirely offline.')
    ap.add_argument('folder', help='folder containing the CVs')
    ap.add_argument('-o', '--out', default='cv_contacts.xlsx',
                    help='output workbook (default: cv_contacts.xlsx)')
    ap.add_argument('--master', help='append clean rows to this existing workbook')
    ap.add_argument('--csv', action='store_true', help='also write a .csv beside the workbook')
    ap.add_argument('--watch', action='store_true',
                    help='keep running and process files as they are added')
    ap.add_argument('--interval', type=int, default=10, help='seconds between polls in --watch')
    args = ap.parse_args()

    if not os.path.isdir(args.folder):
        raise SystemExit('! not a folder: %s' % args.folder)

    records = run_once(args.folder, args.out, args.master)

    if args.csv and records:
        cp = os.path.splitext(args.out)[0] + '.csv'
        with open(cp, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(['Name', 'Email', 'Phone', 'Source File', 'Read As',
                        'Other Emails', 'Flags'])
            for r in records:
                w.writerow([r['name'], r['email'], r['phones'], r['file'],
                            r['how'], r['candidates'], ', '.join(r['flags'])])
        print('csv: %s' % os.path.abspath(cp))

    if args.watch:
        done = {r['file'] for r in records}
        print('\nWatching %s - Ctrl+C to stop.' % args.folder)
        try:
            while True:
                time.sleep(args.interval)
                new = [p for p in collect(args.folder)
                       if os.path.basename(p) not in done]
                if not new:
                    continue
                print('\n%d new file(s)' % len(new))
                for p in new:
                    done.add(os.path.basename(p))
                    rec = process_file(p)
                    records.append(rec)
                    mark = '!' if needs_review(rec) else ' '
                    print('%s %-42s %-24s %-30s %s'
                          % (mark, rec['file'][:42], rec['name'][:24],
                             rec['email'][:30], rec['phones']))
                total, ok, review = write_workbook(records, args.out)
                print('%d CVs | %d clean | %d need review -> %s'
                      % (total, ok, review, args.out))
        except KeyboardInterrupt:
            print('\nStopped.')


if __name__ == '__main__':
    main()
